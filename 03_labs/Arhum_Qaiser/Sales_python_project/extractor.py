from openpyxl import load_workbook
from logger_config import setup_logger
from config import excel_file_path

class extractor : 
    def __init__(self) :
        self.logger = setup_logger()

    def read_excel(self):
        
        data = []
        try:
            workbook = load_workbook(excel_file_path)
            sheet = workbook.active
            headers = [cell.value for cell in sheet[1]]

            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(headers, row))
                data.append(row_data)

            self.logger.info(f"Successfully read {len(data)} rows from Excel.")
        except FileNotFoundError:
            self.logger.error("Excel file not found. Check the file path.")
        except Exception as e:
            self.logger.error(f"Error reading Excel file: {e}")

        return data